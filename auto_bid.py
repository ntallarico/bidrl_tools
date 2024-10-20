import os, sys, getpass
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from config import user_email, user_password
from datetime import datetime
import bidrl_functions as bf
from bidrl_classes import Item, Invoice, Auction
from openpyxl import load_workbook
from ynab_invoice_transaction_split import ynab_invoice_transaction_split_main


# define class Item_AutoBid that inherits all fields from Item class and adds additional fields used for auto bid process
# adds fields related to the concept of an item bid group. These are used in the implementation of the following functionality:
    # the user can specify that multiple items belong to the same group
    # auto_bid will only attempt to win x amount of these items and no more
    # for example: we want to win a bike. 6 bikes are listed on the auction. the user can assign all 6 bikes the same item_bid_group_id
        # and set ibg_items_to_win to 1, and this script will continue to bid on every bike in the group until exactly 1
        # bike has been won by the user's account, and then not bid on any further bikes in the list
class Item_AutoBid(Item):
    def __init__(self
                 , has_autobid_been_placed: int = None
                 , items_in_bid_group: int = None # num of items in the same bid group
                 , items_in_bid_group_won: int = None # num of items in same group that we won / are winning
                 , ibg_items_to_win: int = None # num of items in a bid group that we intend to win
                 #, max_desired_bid: float = None
                 #, item_bid_group_id: str = None
                 , *args, **kwargs):
        
        super().__init__(*args, **kwargs) # call the constructor of the base Item class

        if has_autobid_been_placed is not None and not isinstance(has_autobid_been_placed, int):
            raise TypeError(f"Expected has_autobid_been_placed to be int, got {type(id).__name__}")
        if items_in_bid_group is not None and not isinstance(items_in_bid_group, int):
            raise TypeError(f"Expected items_in_bid_group to be int, got {type(id).__name__}")
        if items_in_bid_group_won is not None and not isinstance(items_in_bid_group_won, int):
            raise TypeError(f"Expected items_in_bid_group_won to be int, got {type(id).__name__}")
        if ibg_items_to_win is not None and not isinstance(ibg_items_to_win, int):
            raise TypeError(f"Expected ibg_items_to_win to be int, got {type(id).__name__}")
        #if max_desired_bid is not None and not isinstance(max_desired_bid, float):
            #raise TypeError(f"Expected max_desired_bid to be float, got {type(max_desired_bid).__name__}")
        #if item_bid_group_id is not None and not isinstance(item_bid_group_id, str):
            #raise TypeError(f"Expected item_bid_group_id to be str, got {type(item_bid_group_id).__name__}")
        
        self.has_autobid_been_placed = has_autobid_been_placed
        self.items_in_bid_group = items_in_bid_group
        self.items_in_bid_group_won = items_in_bid_group_won
        self.ibg_items_to_win = ibg_items_to_win
    
    def display_new_fields(self):
        print(f"Still need to bid?: {self.has_autobid_been_placed}")
        print(f"items_in_bid_group: {self.items_in_bid_group}")
        print(f"items_in_bid_group_won: {self.items_in_bid_group}")
        print(f"ibg_items_to_win: {self.items_in_bid_group}")
        #print(f"Max Desired Bid: {self.max_desired_bid}")
        #print(f"Item Bid Group ID: {self.item_bid_group_id}")


def convert_seconds_to_time_string(seconds):
    days, seconds = divmod(seconds, 86400)  # 60 seconds * 60 minutes * 24 hours
    hours, seconds = divmod(seconds, 3600)  # 60 seconds * 60 minutes
    minutes, seconds = divmod(seconds, 60)
    time_string = ''

    # if days > 0:
    #     time_string += f"{days}d, "
    # if hours > 0:
    #     time_string += f"{hours}h, "
    # if minutes > 0:
    #     time_string += f"{minutes}m, "
    # if seconds >= 0:
    #     time_string += f"{seconds}s"

    if days >= 1:
        time_string += f"{days}d, {hours}h"
    elif hours >= 1:
        time_string += f"{hours}h, {minutes}m"
    else:
        time_string += f"{minutes}m, {seconds}s"

    return time_string



# return current system time in unix format
def time_unix():
    return int(time.time())


# return system format in format "4/29 11:15am"
def time_formatted():
    system_time = datetime.now() # get current system time
    format_1 = system_time.strftime("%m/%d/%y %I:%M%p").lower() # format system time as desired
    format_2 = format_1.lstrip("0").replace("/0", "/") # remove leading zeros
    return format_2


# updates select item info in a list of item objects
# requires: webdriver object, list of item objects
# returns: nothing
# changes: item objects in list
def update_item_info(browser, items):
    try:
        print("Updating item data from bidrl.")

        # get distinct list of auction_ids from items list
        auction_ids = []
        for item in items:
            if item.auction_id not in auction_ids:
                auction_ids.append(item.auction_id)

        new_items = []
        # fast scrape all items from all auctions in list
        for auction_id in auction_ids:
            scraped_items = bf.scrape_items_fast(browser, auction_id, get_images = 'false')
            new_items.extend(scraped_items) # union lists together into new_items
        
        # update fields in items list
        for item in items:
            for new_item in new_items:
                if new_item.id == item.id:
                    item.end_time_unix = new_item.end_time_unix
                    item.highbidder_username = new_item.highbidder_username
                    item.bidding_status = new_item.bidding_status
                    item.current_bid = new_item.current_bid

        print("Success.")
        return 0
    except Exception as e:
        # return 1 here instead of letting this kill the program because update_item_info is not critical to bidding
        print(f"update_item_info() failed with exception: {e}")
        return 1
    

# updates item info pertaining to item bid groups
# loops through list of items, then loops through another list of items for searching
# if search item is in same bid group as our item, then iterate items_in_bid_group
# if search item is in the same bid group as our item and we're winning it, then iterate items_in_bid_group_won
def update_item_group_info(browser, items, username):
    try:
        update_item_info(browser, items)
        print("Updating item group info.")
        for item in items:
            item.items_in_bid_group_won = 0
            if item.highbidder_username == username:
                item.items_in_bid_group_won = 1
            item.items_in_bid_group = 1
            for item_search in items:
                # check if this other item is in the same bid group
                if item.item_bid_group_id == item_search.item_bid_group_id and item.id != item_search.id and item.item_bid_group_id != None and item.item_bid_group_id != '':
                    item.items_in_bid_group += 1
                    # check and see if we already won this other item that is in the same bid group
                    if item_search.highbidder_username == username:
                        item.items_in_bid_group_won += 1
            #print(f"{item.description} | {item.items_in_bid_group} | {item.items_in_bid_group_won}")
        print("Success.")
        return 0
    except Exception as e:
        # return 1 here instead of letting this kill the program because update_item_group_info is not critical to bidding
        print(f"update_item_group_info() failed with exception: {e}")
        return 1


def create_item_objects_from_rows(item_rows_list):
    item_list = [] # list for item objects to return at the end
    for item_row in item_rows_list:
        # check if max_desired_bid is not empty. if it isn't, then convert to float. if it is, then set to None
        max_desired_bid = float(item_row['max_desired_bid']) if item_row['max_desired_bid'] != '' and item_row['max_desired_bid'] != None else None

        # create item object directly in item_list.append
        item_list.append(Item_AutoBid(
            id = item_row['item_id']
            , auction_id = item_row['auction_id']
            , description = item_row['description'].split('", "')[-1].rstrip('")') # pull description out of hyperlink formula
            , end_time_unix = int(item_row['end_time_unix'])
            , max_desired_bid = max_desired_bid
            , item_bid_group_id = item_row['item_bid_group_id']
            , has_autobid_been_placed = 0
            , ibg_items_to_win = int(item_row['ibg_items_to_win'])
            , current_bid = float(0) # set current_bid to 0 so that any checks pass before updating item info, if we were to make any
            , cost_split = item_row['cost_split']
        ))
    return item_list


# read favorite_items_to_input_max_bid.xlsx and return list of item objects for items where max_desired_bid > 0
def read_user_input_xlsx_to_item_objects(browser, auto_bid_folder_path):
    try:
        filename = 'favorite_items_to_input_max_bid.xlsx'

        file_path = auto_bid_folder_path + filename

        wb = load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]] # get headers from the first row

        read_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            read_rows.append(row_dict)

        print(f"\nRead {len(read_rows)} rows from file: {filename}.")

        item_list = create_item_objects_from_rows(read_rows)
        print(f"Created {len(item_list)} item objects from read rows.")

        # sort list of items in descending order based on their end time
        item_list.sort(key=lambda x: x.end_time_unix, reverse=True)

        return item_list
    except Exception as e:
        print(f"read_user_input_xlsx_to_item_objects() failed with exception: {e}")
        print("Tearing down web object.")
        browser.quit()
        return 1


def get_username(browser):
    username = bf.get_session(browser)['user_name']
    return username


def print_items_status(item_list):
    print('\n----------------------------------------------------------------------------------------------------')
    current_time_unix = time_unix()
    for item in item_list:
        if is_item_eligible_for_bidding(item):
            remaining_seconds = item.end_time_unix - current_time_unix
            remaining_time_string = convert_seconds_to_time_string(remaining_seconds)
            length_formatted_remaining_time_string = f"{remaining_time_string:<8}"
            if item.items_in_bid_group_won >= item.ibg_items_to_win:
                length_formatted_max_desired_bid = 'BG_Done' # display indication that we have already won the # of items desired from the bid group
            elif item.max_desired_bid <= item.current_bid:
                length_formatted_max_desired_bid = 'LOST   '
            else:
                length_formatted_max_desired_bid = f"${str(item.max_desired_bid):<6}"

            if len(item.description) > 63:
                length_formatted_description = item.description[:30] + '...' + item.description[-30:]
            else:
                length_formatted_description = f"{str(item.description):<63}"
            #print(f"{length_formatted_remaining_time_string} | {length_formatted_max_desired_bid} | {length_formatted_description} | {item.ibg_items_to_win}")
            print(f"{length_formatted_description} | {length_formatted_remaining_time_string} | {length_formatted_max_desired_bid} | Group: {item.item_bid_group_id} | Qty desired: {item.ibg_items_to_win}")
    print('----------------------------------------------------------------------------------------------------')


# returns True if item is eligible for bidding. checks if:
    # item is not already closed
    # max_desired_bid is not None
    # max_desired_bid is not 0
    # max_desired_bid > current highest bid on the item
    # auto_bid() has not already placed a bid on this item
# keep this function just referencing local stuff (no calls to bidrl). we run this rapidly and often
def is_item_eligible_for_bidding(item):
    if item.bidding_status != 'Closed' \
        and item.max_desired_bid != None \
        and item.max_desired_bid != 0 \
        and item.has_autobid_been_placed == 0:
        return True
    else:
        return False


# loop through each item and attempt to place a bid on it if we determine that we should
# checks to see if time remaining on the item is <= our set seconds_before_closing_to_bid time and is_item_eligible_for_bidding(item) == True
# if it is eligible and time to bid, then update the item/group info
    # and check if more than [item.ibg_items_to_win] items in its bid group have not already been won
def auto_bid(browser, item_list, seconds_before_closing_to_bid, username):
    current_time_unix = time_unix()
    # loop through each item and bid on it if the time remaining on the item is <= our set seconds_before_closing_to_bid time
    for item in item_list:
        remaining_seconds = item.end_time_unix - current_time_unix
        if remaining_seconds <= seconds_before_closing_to_bid and is_item_eligible_for_bidding(item) and item.items_in_bid_group_won < item.ibg_items_to_win:
            print(f"{remaining_seconds} seconds remaining. Time to pace bid on: {item.description}.")
            update_item_group_info(browser, item_list, username)
            if item.items_in_bid_group_won < item.ibg_items_to_win:
                bf.bid_on_item(item, item.max_desired_bid, browser)
                print('')
                item.has_autobid_been_placed = 1
            else:
                print(f"Already won {item.items_in_bid_group_won} items in bid group!")

# return webdriver instance configured for use in auto_bid operation
# headless, with special identifier "auto_bid" so that we can monitor instances in auto_bid_orchestrator.py
def get_logged_in_webdriver_for_auto_bid():
    return bf.get_logged_in_webdriver(user_email, user_password, headless = 'headless', webdriver_identifier = 'auto_bid')

def login_refresh(browser, last_login_time, last_login_time_unix):
    try:
        print(f"Checking login status ({time_formatted()}).")
        if bf.check_if_login_success(browser) != 0:
            print("Login check failed!")
            print("Tearing down webdriver, initiating new webdriver, and starting login process.")
            browser.quit()
            browser = get_logged_in_webdriver_for_auto_bid()
            last_login_time = time_formatted()
            last_login_time_unix = time_unix()
        else:
            logged_in_time = convert_seconds_to_time_string(time_unix() - last_login_time_unix)
            print(f"Login check success. Last login: {last_login_time}. Time remained logged in: {logged_in_time}.")
    except Exception as e:
        print(f"Exception occurred in login_refresh().\nException: {e}")
        bf.tear_down(browser)

# save information from a list of item objects to the items_user_input table in the database
def update_items_user_input_table(item_list):
    try:
        print("\nAttempting to use information read from csv to update items_user_input table in database.")
        conn = bf.init_sqlite_connection(path = 'local_files/auto_bid/', database = 'bidrl_user_input')
        cursor = conn.cursor()

        for item in item_list:
            if item.max_desired_bid not in [None, 0]:
                cursor.execute("""
                    SELECT COUNT(*) FROM items_user_input WHERE item_id = ? AND auction_id = ?
                """, (item.id, item.auction_id))
                exists = cursor.fetchone()[0]

                # define field names to update in the database
                field_names = ['description', 'url', 'end_time_unix', 'item_bid_group_id', 'ibg_items_to_win', 'max_desired_bid', 'cost_split']

                if exists: # if the item exists in the items_user_input db table, then update its fields based on contents of csv
                    cursor.execute(f"""
                        UPDATE items_user_input
                        SET {', '.join([f"{field} = ?" for field in field_names])}
                        WHERE item_id = ?
                    """, tuple(getattr(item, field) for field in field_names) + (item.id,))
                else: # if the item does not exist in the items_user_input db table, then create it using fields in csv
                    cursor.execute(f"""
                        INSERT INTO items_user_input (item_id, auction_id, {', '.join(field_names)})
                        VALUES ({', '.join(['?' for _ in range(len(field_names) + 2)])})
                    """, (item.id, item.auction_id) + tuple(getattr(item, field) for field in field_names))

        conn.commit()
        print("Closing sqlite connection")
        conn.close()
        print("Success.")
    except Exception as e:
        print(f"Exception occurred in update_items_user_input_table(). Moving forward anyway as this isn't critical for auto_bid.\nException: {e}")

# if ynab_api_token and item up for bid next is not too close, run ynab_invoice_transaction_split_main() to split any uncategorized
# bidrl ynab transactions. by default, will not run if next item is up for bidding within 30 mins
def ynab_trans_split(item_list, desired_distance_from_next_bid_time = 30 * 60):
    try:
        # first check if ynab stuff is defined in config file. if it isn't, then pass
        try:
            from config import ynab_api_token
        except ImportError:
            ynab_api_token = None

        if not ynab_api_token:
            #print("\nynab_api_token is not defined in config.py. Skipping YNAB transaction split.")
            return
        else:
            print("\nynab_api_token found in config.py. proceeding with ynab_trans_split()")

        # look at all items and find the number of seconds until the next item is up for bid
        print("Checking time until next bid")
        current_time_unix = time_unix()
        min_remaining_seconds = float('inf')
        for item in item_list:
            remaining_seconds = item.end_time_unix - current_time_unix
            if is_item_eligible_for_bidding(item) and item.items_in_bid_group_won < item.ibg_items_to_win:
                if remaining_seconds < min_remaining_seconds:
                    min_remaining_seconds = remaining_seconds
        print(f"Remaining seconds: {min_remaining_seconds}")
        
        if min_remaining_seconds > desired_distance_from_next_bid_time:
            print(f"Proceeding with ynab_invoice_transaction_split_main()")
            ynab_invoice_transaction_split_main()
        else:
            print(f"Skipping ynab_invoice_transaction_split_main()")
            return
    except Exception as e:
        print(f"ynab_trans_split() failed with exception: {e}")
        return

def auto_bid_main(seconds_before_closing_to_bid = 120 + 5 # add 5 secs to account for POST time to API. don't want to extend bid time if we can avoid
         , auto_bid__interval = 5 # how often to check if it is time to bid on each item (and then bid if it is)
         , print_items_status__interval = 60 # how often to print times remaining for all items
         , login_refresh__interval = 60 # keep this reasonable - actually submits a request to bidrl
         , update_item_info__interval = 60 * 60 # keep this reasonable - actually submits a request to bidrl
         , ynab_trans_split__interval = 60 * 30 # how often to check if ynab has any uncategorized transactions to split
         ):

    browser = None # establish browser variable so that the check in 'finally' doesn't throw an error

    try:
        auto_bid_folder_path = 'local_files/auto_bid/'
        bf.ensure_directory_exists(auto_bid_folder_path)
        
        # get an initialized web driver that has logged in to bidrl with credentials stored in config.py
        browser = get_logged_in_webdriver_for_auto_bid()
        last_login_time_string = time_formatted()
        last_login_time_unix = time_unix()
        username = get_username(browser)
        print(f"Username: {username}")

        # read favorite_items_to_input_max_bid.csv and return list of item objects we intend to bid on
        item_list = read_user_input_xlsx_to_item_objects(browser, auto_bid_folder_path)

        update_item_group_info(browser, item_list, username)

        # save inputs from our csv to the database for data archive purposes
        update_items_user_input_table(item_list)

        item_count_with_desired_bid = 0
        item_count_zero_desired_bid = 0
        item_count_no_desired_bid = 0
        item_count_closed = 0
        item_count_actually_intend_to_bid = 0
        for item in item_list:
            if item.max_desired_bid == None:
                item_count_no_desired_bid += 1
            elif item.max_desired_bid == 0:
                item_count_zero_desired_bid += 1
            else: # item has a max_desired_bid
                item_count_with_desired_bid += 1
                if item.bidding_status == 'Closed':
                    item_count_closed += 1
                else:
                    item_count_actually_intend_to_bid += 1

        print(f"\nItems without max_desired_bid: {item_count_no_desired_bid}")
        print(f"Items with 0 max_desired_bid: {item_count_zero_desired_bid}")
        print(f"Items with max_desired_bid: {item_count_with_desired_bid}")
        print(f"Items with max_desired_bid that already closed: {item_count_closed}")
        print(f"\nWe intend to bid on {item_count_actually_intend_to_bid} items, {seconds_before_closing_to_bid} seconds before they close, checking every {auto_bid__interval} seconds.\n")

        # set x__last_run_time to 0 to run immediately when loop processes
        # set to time_unix() to wait x__interval amount of seconds first
        auto_bid__last_run_time = 0
        login_refresh__last_run_time = 0
        print_items_status__last_run_time = 0
        update_item_info__last_run_time = time_unix()
        ynab_trans_split__last_run_time = 0

        try:
            # loop until KeyboardInterrupt, testing if it has been x_function__interval seconds since last run of x_function()
            # if it has been, run x_function(), then test the next function and its corresponding times
            while True:
                # update_item_info()
                if time_unix() - update_item_info__last_run_time >= update_item_info__interval:
                    update_item_info(browser, item_list)
                    update_item_info__last_run_time = time_unix()

                # login_refresh()
                if time_unix() - login_refresh__last_run_time >= login_refresh__interval:
                    login_refresh(browser, last_login_time_string, last_login_time_unix)
                    login_refresh__last_run_time = time_unix()

                # print_items_status()
                if time_unix() - print_items_status__last_run_time >= print_items_status__interval:
                    print_items_status(item_list)
                    print_items_status__last_run_time = time_unix()

                # auto_bid()
                if time_unix() - auto_bid__last_run_time >= auto_bid__interval:
                    auto_bid(browser, item_list, seconds_before_closing_to_bid, username)
                    auto_bid__last_run_time = time_unix()

                # ynab_trans_split()
                if time_unix() - ynab_trans_split__last_run_time >= ynab_trans_split__interval:
                    ynab_trans_split(item_list)
                    ynab_trans_split__last_run_time = time_unix()

                time.sleep(1)
        finally:
            print("Loop interrupted.")
            if browser: bf.tear_down(browser)
    except Exception as e:
        print(f"Exception occurred in auto_bid_main().\nException: {e}")
        if browser: bf.tear_down(browser)
    finally:
        if browser: bf.tear_down(browser)


if __name__ == "__main__":
    auto_bid_main()

