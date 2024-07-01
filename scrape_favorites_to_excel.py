import os
import csv
from config import user_email, user_password, home_affiliates
from datetime import datetime
import bidrl_functions as bf
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule

auto_bid_folder_path = 'local_files/auto_bid/'
bf.ensure_directory_exists(auto_bid_folder_path)

filename_to_write = auto_bid_folder_path + 'favorite_items_to_input_max_bid.xlsx'

# check if the file already exists and ask for user input to confirm overwriting it if it does
if os.path.exists(filename_to_write):
    user_input = input(f"\nThe file '{filename_to_write}' already exists, possibly with your max bids in it. Do you want to overwrite it? (y/n): ")
    
    if user_input.lower() != 'y':
        print("File will not be overwritten. Exiting.")
        quit()
    else:
        print(f"Will overwrite '{filename_to_write}'. Deleting the existing file.")
        os.remove(filename_to_write)
else:
    print(f"Creating '{filename_to_write}'.")


# get an initialized web driver that has logged in to bidrl with credentials stored in config.py
browser = bf.get_logged_in_webdriver(user_email, user_password, 'headless')


# use get_open_auctions_fast() to get abridged version of all open auction/item data for our home affiliates
open_auctions = []
for aff in home_affiliates:
    open_auctions.extend(bf.get_open_auctions_fast(browser, affiliate_id = aff))


print(f"\nFinding favorites from {len(open_auctions)} auctions.")
rows_to_write = []
for auction in open_auctions:
    for item in auction.items:
        if item.is_favorite == 1:
            try:
                rows_to_write.append({
                    'end_time_unix': item.end_time_unix
                    , 'auction_id': auction.id
                    , 'item_id': item.id
                    , 'description': f"=HYPERLINK(\"{item.url}\", \"{item.description}\")" # create as hyperlink to item page
                    , 'max_desired_bid': ''  # placeholder for user input
                    , 'cost_split': '' # placeholder for user input
                    , 'item_bid_group_id': item.id
                    , 'ibg_items_to_win': 1
                })
                print(f"Found favorite: {item.description}")
            except Exception as e:
                print(f"Failed on item: {item.description}")
                print(f"Exception: {e}")
                bf.tear_down(browser)
print(f"\nFound {len(rows_to_write)} favorites.")


try:
    wb = Workbook()
    ws = wb.active
    ws.title = "Favorite Items"

    # Write column headers
    if rows_to_write:
        ws.append(list(rows_to_write[0].keys()))

    # Write item data
    for row in rows_to_write:
        ws.append(list(row.values()))

    num_of_columns = len(rows_to_write[0].keys())

    # Create a table with a specific style
    right_most_column_letter = chr(64 + num_of_columns)
    tab = Table(displayName="FavoritesTable", ref=f"A1:{right_most_column_letter}{len(rows_to_write) + 1}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Apply alternating row colors for better readability
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(rows_to_write) + 1, min_col=1, max_col=num_of_columns), start=2):
        fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") if idx % 2 == 0 else PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for cell in row:
            cell.fill = fill


    # find column letters based on their names for reference in formulas
    max_desired_bid_col = None
    cost_split_col = None
    description_col = None
    for cell in ws[1]:  # Iterate over the first row to find the column names
        if cell.value == 'max_desired_bid':
            max_desired_bid_col = cell.column_letter
        elif cell.value == 'cost_split':
            cost_split_col = cell.column_letter
        elif cell.value == 'description':
            description_col = cell.column_letter


    # right-justify description column
    if description_col:
        for cell in ws[description_col]:
            cell.alignment = Alignment(horizontal='right')
    

    # Highlight cells in the 'max_desired_bid' column if they are empty
    if max_desired_bid_col and cost_split_col:
        ws.conditional_formatting.add(f'{max_desired_bid_col}2:{max_desired_bid_col}{len(rows_to_write) + 1}',
                                      CellIsRule(operator='equal', formula=['""'], stopIfTrue=True, fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")))
        
    # Highlight cells in the 'cost_split' column if they are empty and 'max_desired_bid' is not zero
    formula = f'AND(ISBLANK({cost_split_col}2), {max_desired_bid_col}2<>0)'
    ws.conditional_formatting.add(f'{cost_split_col}2:{cost_split_col}{len(rows_to_write) + 1}',
                                  FormulaRule(formula=[formula], fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")))


    # Freeze the first row of the table
    ws.freeze_panes = ws['A2']

    # Set column widths
    column_widths = {
        "end_time_unix": 10.3,
        "auction_id": 6.7,
        "item_id": 8.8,
        "item_bid_group_id": 19.7,
        "ibg_items_to_win": 18.6,
        "description": 100,
        "max_desired_bid": 8,
        "cost_split": 4
    }

    # unsure why, but they come out as aproximately 0.75 smaller than I input, so adding that amount to my indended inputs
    for key in column_widths:
        column_widths[key] += 0.75

    for col_name, width in column_widths.items():
        for cell in ws[1]:  # Iterate over the first row to find the column names
            if cell.value == col_name:
                ws.column_dimensions[cell.column_letter].width = width
                break


    wb.save(filename_to_write)
    print(f"\nFile successfully written: {filename_to_write}")
except Exception as e:
    print(f"Attempt to write to xlsx failed.")
    print(f"Exception: {e}")
    bf.tear_down(browser)

bf.tear_down(browser)