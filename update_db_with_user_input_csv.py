import os, sys, getpass
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from config import user_email, user_password
from datetime import datetime
import bidrl_functions as bf
from bidrl_classes import Item, Invoice, Auction
from openpyxl import load_workbook
from ynab_invoice_transaction_split import ynab_invoice_transaction_split_main


# define class Item_AutoBid that inherits all fields from Item class and adds additional fields used for auto bid process
# adds fields related to the concept of an item bid group. These are used in the implementation of the following functionality:
    # the user can specify that multiple items belong to the same group
    # auto_bid will only attempt to win x amount of these items and no more
    # for example: we want to win a bike. 6 bikes are listed on the auction. the user can assign all 6 bikes the same item_bid_group_id
        # and set ibg_items_to_win to 1, and this script will continue to bid on every bike in the group until exactly 1
        # bike has been won by the user's account, and then not bid on any further bikes in the list
class Item_AutoBid(Item):
    def __init__(self
                 , has_autobid_been_placed: int = None
                 , items_in_bid_group: int = None # num of items in the same bid group
                 , items_in_bid_group_won: int = None # num of items in same group that we won / are winning
                 , ibg_items_to_win: int = None # num of items in a bid group that we intend to win
                 #, max_desired_bid: float = None
                 #, item_bid_group_id: str = None
                 , *args, **kwargs):
        
        super().__init__(*args, **kwargs) # call the constructor of the base Item class

        if has_autobid_been_placed is not None and not isinstance(has_autobid_been_placed, int):
            raise TypeError(f"Expected has_autobid_been_placed to be int, got {type(id).__name__}")
        if items_in_bid_group is not None and not isinstance(items_in_bid_group, int):
            raise TypeError(f"Expected items_in_bid_group to be int, got {type(id).__name__}")
        if items_in_bid_group_won is not None and not isinstance(items_in_bid_group_won, int):
            raise TypeError(f"Expected items_in_bid_group_won to be int, got {type(id).__name__}")
        if ibg_items_to_win is not None and not isinstance(ibg_items_to_win, int):
            raise TypeError(f"Expected ibg_items_to_win to be int, got {type(id).__name__}")
        #if max_desired_bid is not None and not isinstance(max_desired_bid, float):
            #raise TypeError(f"Expected max_desired_bid to be float, got {type(max_desired_bid).__name__}")
        #if item_bid_group_id is not None and not isinstance(item_bid_group_id, str):
            #raise TypeError(f"Expected item_bid_group_id to be str, got {type(item_bid_group_id).__name__}")
        
        self.has_autobid_been_placed = has_autobid_been_placed
        self.items_in_bid_group = items_in_bid_group
        self.items_in_bid_group_won = items_in_bid_group_won
        self.ibg_items_to_win = ibg_items_to_win
    
    def display_new_fields(self):
        print(f"Still need to bid?: {self.has_autobid_been_placed}")
        print(f"items_in_bid_group: {self.items_in_bid_group}")
        print(f"items_in_bid_group_won: {self.items_in_bid_group}")
        print(f"ibg_items_to_win: {self.items_in_bid_group}")
        #print(f"Max Desired Bid: {self.max_desired_bid}")
        #print(f"Item Bid Group ID: {self.item_bid_group_id}")


# return current system time in unix format
def time_unix():
    return int(time.time())

# updates select item info in a list of item objects
# requires: webdriver object, list of item objects
# returns: nothing
# changes: item objects in list
def update_item_info(browser, items):
    try:
        print("Updating item data from bidrl.")

        # get distinct list of auction_ids from items list
        auction_ids = []
        for item in items:
            if item.auction_id not in auction_ids:
                auction_ids.append(item.auction_id)

        new_items = []
        # fast scrape all items from all auctions in list
        for auction_id in auction_ids:
            scraped_items = bf.scrape_items_fast(browser, auction_id, get_images = 'false')
            new_items.extend(scraped_items) # union lists together into new_items
        
        # update fields in items list
        for item in items:
            for new_item in new_items:
                if new_item.id == item.id:
                    item.end_time_unix = new_item.end_time_unix
                    item.highbidder_username = new_item.highbidder_username
                    item.bidding_status = new_item.bidding_status
                    item.current_bid = new_item.current_bid
                    item.url = new_item.url

        print("Success.")
        return 0
    except Exception as e:
        # return 1 here instead of letting this kill the program because update_item_info is not critical to bidding
        print(f"update_item_info() failed with exception: {e}")
        return 1
    

# updates item info pertaining to item bid groups
# loops through list of items, then loops through another list of items for searching
# if search item is in same bid group as our item, then iterate items_in_bid_group
# if search item is in the same bid group as our item and we're winning it, then iterate items_in_bid_group_won
def update_item_group_info(browser, items, username):
    try:
        update_item_info(browser, items)
        print("Updating item group info.")
        for item in items:
            item.items_in_bid_group_won = 0
            if item.highbidder_username == username:
                item.items_in_bid_group_won = 1
            item.items_in_bid_group = 1
            for item_search in items:
                # check if this other item is in the same bid group
                if item.item_bid_group_id == item_search.item_bid_group_id and item.id != item_search.id and item.item_bid_group_id != None and item.item_bid_group_id != '':
                    item.items_in_bid_group += 1
                    # check and see if we already won this other item that is in the same bid group
                    if item_search.highbidder_username == username:
                        item.items_in_bid_group_won += 1
            #print(f"{item.description} | {item.items_in_bid_group} | {item.items_in_bid_group_won}")
        print("Success.")
        return 0
    except Exception as e:
        # return 1 here instead of letting this kill the program because update_item_group_info is not critical to bidding
        print(f"update_item_group_info() failed with exception: {e}")
        return 1


def create_item_objects_from_rows(item_rows_list):
    item_list = [] # list for item objects to return at the end
    for item_row in item_rows_list:
        # check if max_desired_bid is not empty. if it isn't, then convert to float. if it is, then set to None
        max_desired_bid = float(item_row['max_desired_bid']) if item_row['max_desired_bid'] != '' and item_row['max_desired_bid'] != None else None

        # create item object directly in item_list.append
        item_list.append(Item_AutoBid(
            id = item_row['item_id']
            , auction_id = item_row['auction_id']
            , description = item_row['description'].split('", "')[-1].rstrip('")') # pull description out of hyperlink formula
            , end_time_unix = int(item_row['end_time_unix'])
            , max_desired_bid = max_desired_bid
            , item_bid_group_id = item_row['item_bid_group_id']
            , has_autobid_been_placed = 0
            , ibg_items_to_win = int(item_row['ibg_items_to_win'])
            , current_bid = float(0) # set current_bid to 0 so that any checks pass before updating item info, if we were to make any
            , cost_split = item_row['cost_split']
        ))
    return item_list


# read favorite_items_to_input_max_bid.xlsx and return list of item objects for items where max_desired_bid > 0
def read_user_input_xlsx_to_item_objects(browser, auto_bid_folder_path):
    try:
        filename = 'favorite_items_to_input_max_bid.xlsx'

        file_path = auto_bid_folder_path + filename

        wb = load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]] # get headers from the first row

        read_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            read_rows.append(row_dict)

        print(f"\nRead {len(read_rows)} rows from file: {filename}.")

        item_list = create_item_objects_from_rows(read_rows)
        print(f"Created {len(item_list)} item objects from read rows.")

        # sort list of items in descending order based on their end time
        item_list.sort(key=lambda x: x.end_time_unix, reverse=True)

        return item_list
    except Exception as e:
        print(f"read_user_input_xlsx_to_item_objects() failed with exception: {e}")
        print("Tearing down web object.")
        browser.quit()
        return 1

# return webdriver instance configured for use in auto_bid operation
# headless, with special identifier "auto_bid" so that we can monitor instances in auto_bid_orchestrator.py
def get_logged_in_webdriver_for_auto_bid():
    return bf.get_logged_in_webdriver(user_email, user_password, headless = 'headless', webdriver_identifier = 'update_db_with_user_input_csv')

# save information from a list of item objects to the auto_bid_itemuserinput table in the database
def update_items_user_input_table(item_list):
    try:
        print("\nAttempting to use information read from csv to update auto_bid_itemuserinput table in database.")
        conn = bf.init_sqlite_connection(path = 'local_files/auto_bid/', database = 'auto_bid')
        cursor = conn.cursor()

        for item in item_list:
            if item.max_desired_bid not in [None, 0]:
                cursor.execute("""
                    SELECT COUNT(*) FROM auto_bid_itemuserinput WHERE item_id = ? AND auction_id = ?
                """, (item.id, item.auction_id))
                exists = cursor.fetchone()[0]

                # define field names to update in the database
                field_names = ['description', 'url', 'end_time_unix', 'item_bid_group_id', 'ibg_items_to_win', 'max_desired_bid', 'cost_split']

                if exists: # if the item exists in the auto_bid_itemuserinput db table, then update its fields based on contents of csv
                    cursor.execute(f"""
                        UPDATE auto_bid_itemuserinput
                        SET {', '.join([f"{field} = ?" for field in field_names])}
                        WHERE item_id = ?
                    """, tuple(getattr(item, field) for field in field_names) + (item.id,))
                else: # if the item does not exist in the auto_bid_itemuserinput db table, then create it using fields in csv
                    cursor.execute(f"""
                        INSERT INTO auto_bid_itemuserinput (item_id, auction_id, {', '.join(field_names)})
                        VALUES ({', '.join(['?' for _ in range(len(field_names) + 2)])})
                    """, (item.id, item.auction_id) + tuple(getattr(item, field) for field in field_names))

        conn.commit()
        print("Closing sqlite connection")
        conn.close()
        print("Success.")
    except Exception as e:
        print(f"Exception occurred in update_items_user_input_table(): {e}")

def get_username(browser):
    username = bf.get_session(browser)['user_name']
    return username

def update_db_with_user_input_csv_main():

    browser = None # establish browser variable so that the check in 'finally' doesn't throw an error

    try:
        auto_bid_folder_path = 'local_files/auto_bid/'
        bf.ensure_directory_exists(auto_bid_folder_path)
        
        # get an initialized web driver that has logged in to bidrl with credentials stored in config.py
        browser = get_logged_in_webdriver_for_auto_bid()

        username = get_username(browser)
        print(f"Username: {username}")

        # read favorite_items_to_input_max_bid.csv and return list of item objects we intend to bid on
        item_list = read_user_input_xlsx_to_item_objects(browser, auto_bid_folder_path)

        update_item_group_info(browser, item_list, username)

        # save inputs from our csv to the database for data archive purposes
        update_items_user_input_table(item_list)

        item_count_with_desired_bid = 0
        item_count_zero_desired_bid = 0
        item_count_no_desired_bid = 0
        item_count_closed = 0
        item_count_actually_intend_to_bid = 0
        for item in item_list:
            if item.max_desired_bid == None:
                item_count_no_desired_bid += 1
            elif item.max_desired_bid == 0:
                item_count_zero_desired_bid += 1
            else: # item has a max_desired_bid
                item_count_with_desired_bid += 1
                if item.bidding_status == 'Closed':
                    item_count_closed += 1
                else:
                    item_count_actually_intend_to_bid += 1

        print(f"\nItems without max_desired_bid: {item_count_no_desired_bid}")
        print(f"Items with 0 max_desired_bid: {item_count_zero_desired_bid}")
        print(f"Items with max_desired_bid: {item_count_with_desired_bid}")
        print(f"Items with max_desired_bid that already closed: {item_count_closed}")
    except Exception as e:
        print(f"Exception occurred in auto_bid_main().\nException: {e}")
        if browser: bf.tear_down(browser)
    finally:
        if browser: bf.tear_down(browser)


if __name__ == "__main__":
    update_db_with_user_input_csv_main()

